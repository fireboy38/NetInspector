#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel/CSV 设备清单解析器
"""

import os
import logging

logger = logging.getLogger(__name__)

# 表头映射 (支持中英文)
COLUMN_MAP = {
    'name': ['设备名称', 'name', 'hostname', '名称', 'device_name'],
    'platform': ['设备类型', 'platform', 'vendor', '厂商', 'type', '类型'],
    'host': ['ip地址', 'ip', 'host', 'address', '地址', 'ip_address'],
    'port': ['端口号', 'port', '端口'],
    'protocol': ['连接类型', 'protocol', '协议', 'connection_type'],
    'username': ['用户名', 'username', 'user', 'login'],
    'password': ['密码', 'password', 'passwd', 'pass'],
    'enable_password': ['特权密码', 'enable', 'enable_password', 'secret', '使能密码'],
}


def _find_col(headers: list, field: str) -> int:
    """在表头中查找字段索引，不区分大小写"""
    candidates = COLUMN_MAP.get(field, [field])
    for i, h in enumerate(headers):
        if h.strip().lower() in [c.lower() for c in candidates]:
            return i
    return -1


def parse_excel(filepath: str) -> list:
    """
    解析 Excel 文件，返回设备信息列表
    每条记录为 dict: name/platform/host/port/protocol/username/password/enable_password
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return _parse_excel_openpyxl(filepath)
    elif ext == '.csv':
        return _parse_csv(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx/.xls/.csv")


def _parse_excel_openpyxl(filepath: str) -> list:
    """使用 openpyxl 解析 xlsx 文件"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    return _parse_rows(rows)


def _parse_csv(filepath: str) -> list:
    """解析 CSV 文件"""
    import csv
    rows = []
    for enc in ('utf-8-sig', 'gbk', 'utf-8'):
        try:
            with open(filepath, encoding=enc, newline='') as f:
                reader = csv.reader(f)
                rows = [tuple(r) for r in reader]
            break
        except UnicodeDecodeError:
            continue
    return _parse_rows(rows)


def _parse_rows(rows: list) -> list:
    """通用行数据解析"""
    # 找表头行（第一行）
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]

    col_name = _find_col(headers, 'name')
    col_platform = _find_col(headers, 'platform')
    col_host = _find_col(headers, 'host')
    col_port = _find_col(headers, 'port')
    col_proto = _find_col(headers, 'protocol')
    col_user = _find_col(headers, 'username')
    col_pass = _find_col(headers, 'password')
    col_enable = _find_col(headers, 'enable_password')

    devices = []
    for row in rows[1:]:
        if not any(row):
            continue

        def get(idx, default=''):
            if idx < 0 or idx >= len(row):
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        host = get(col_host)
        if not host:
            continue  # 必须有IP

        port_str = get(col_port, '')
        proto = get(col_proto, 'ssh').lower()
        if not port_str:
            port = 22 if 'ssh' in proto else 23
        else:
            try:
                port = int(float(port_str))
            except Exception:
                port = 22 if 'ssh' in proto else 23

        devices.append({
            'name': get(col_name, host),
            'platform': get(col_platform, 'default').lower(),
            'host': host,
            'port': port,
            'protocol': proto,
            'username': get(col_user, 'admin'),
            'password': get(col_pass, ''),
            'enable_password': get(col_enable, ''),
        })
    return devices


def generate_template(filepath: str):
    """生成设备清单模板 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设备清单'

    headers = ['设备名称', '设备类型', 'IP地址', '端口号', '连接类型', '用户名', '密码', '特权密码']
    sample_data = [
        ['华为核心交换机', 'huawei', '192.168.1.1', '22', 'ssh', 'admin', 'Admin@123', ''],
        ['H3C汇聚交换机', 'h3c', '192.168.1.2', '22', 'ssh', 'admin', 'Admin@123', ''],
        ['思科交换机', 'cisco', '192.168.1.3', '22', 'ssh', 'admin', 'Admin@123', 'enable123'],
        ['锐捷交换机', 'ruijie', '192.168.1.4', '23', 'telnet', 'admin', 'Admin@123', ''],
    ]

    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [20, 12, 18, 8, 12, 12, 15, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    wb.save(filepath)
    return filepath
